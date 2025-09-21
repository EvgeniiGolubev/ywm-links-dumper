import os
import requests
from pathlib import Path
from typing import Dict, Optional, Iterable
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

API_BASE = "https://api.webmaster.yandex.net/v4"
XLSX_FIELDS = ["source_url", "destination_url", "discovery_date", "source_last_access_date"]
BAD_KEYWORDS = [
    # adult
    "porn", "porno", "pornhub", "xvideos", "xnxx", "xxx",
    "sex", "hentai", "nude", "cam", "cams", "webcam",
    "onlyfans", "erotic", "escort",

    # gambling
    "casino", "slots", "slot", "roulette", "poker",
    "bookmaker", "bookie", "bet", "bets", "betting",
    "stake", "wager", "1xbet", "parimatch", "fonbet",
    "leonbets", "melbet", "vulkan", "joycasino",
    "pin-up", "pinup", "gg.bet", "bet365", "marathonbet",

    # loans / scams
    "payday", "loan", "microloan", "mfo",

    # pharma
    "viagra", "cialis", "levitra", "tramadol", "modafinil",
    "steroids", "anabolic", "clenbuterol", "pharmacy", "pills",

    # drugs
    "cannabis", "marijuana", "weed", "hash", "kush", "cbd",
    "lsd", "mdma", "ecstasy",

    # piracy
    "torrent", "torrents", "warez", "pirate", "crack",
    "keygen", "serials", "nulled", "mod apk", "apk mod",
    "free download full",

    # dating / spam
    "adult", "dating", "hookup", "milfs", "camgirls",
    "livejasmin", "bongacams", "chaturbate",

    # hyip / binary options
    "hyip", "ponzi", "binary", "options",
]

def make_headers(token: str) -> Dict[str, str]:
    return {"Authorization": f"OAuth {token}", "Accept": "application/json"}

def get_user_id(headers: Dict[str, str]) -> int:
    r = requests.get(f"{API_BASE}/user", headers=headers, timeout=60)
    r.raise_for_status() # Если ошибка клиента или сервера кидает исключение HTTPError
    data = r.json()
    uid = data.get("user_id")

    if not uid:
        raise RuntimeError(f"Не удалось получить user_id из ответа: {data}")

    return int(uid)

def find_host_id(headers: Dict[str, str], user_id: int, host_domain: str) -> Optional[str]:
    r = requests.get(f"{API_BASE}/user/{user_id}/hosts", headers=headers, timeout=60)
    r.raise_for_status()
    data = r.json()
    hosts = data.get("hosts")

    for h in hosts:
        hid = h.get("host_id")

        if not hid:
            continue

        if host_domain in hid:
            return hid

    return None

def iter_external_links(headers: Dict[str, str], user_id: int, host_id: str, limit: int, offset: int) -> Iterable[Dict]:
    while True:
        url = f"{API_BASE}/user/{user_id}/hosts/{host_id}/links/external/samples"
        params = {"limit": limit, "offset": offset}
        resp = requests.get(url, headers=headers, params=params, timeout=90)

        if resp.status_code == 401:
            raise RuntimeError("401 Unauthorized: проверьте YWM_OAUTH_TOKEN")
        if resp.status_code == 403:
            raise RuntimeError("403 Forbidden: нет прав на этот host_id")

        resp.raise_for_status()

        data = resp.json()
        links = data.get("links", [])

        for link in links:
            yield link

        got = len(links)

        print(f"Получено {got} ссылок, offset={offset}")

        if got < limit:
            break

        offset += limit

def make_file_name(host_domain):
    today = datetime.today().strftime("%Y-%m-%d")
    safe_domain = host_domain.replace(":", "_")
    return f"{safe_domain}_links_{today}.xlsx"

def is_bad_url(url: str) -> bool:
    s = (url or "").lower()
    for k in BAD_KEYWORDS:
        if k in s:
            print(k)

    return any(k in s for k in BAD_KEYWORDS)

def write_xlsx(rows: Iterable[Dict], out_dir: Path, host_domain: str) -> None:
    file_name = make_file_name(host_domain)
    out_path = out_dir / file_name

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "External Links"

    ws.append(XLSX_FIELDS)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for row in rows:
        values = [row.get(k, "") for k in XLSX_FIELDS]
        ws.append(values)

    # подсветка "плохих" ссылок
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        source_url = str(row[0].value).lower()
        if is_bad_url(source_url):
            for cell in row:
                cell.fill = red_fill

    wb.save(out_path)

    print(f"Файл сохранён: {out_path.resolve()}")

def main() -> None:
    load_dotenv()

    token = os.getenv("YWM_OAUTH_TOKEN")
    host_domain = os.getenv("YWM_HOST_DOMAIN")
    limit = int(os.getenv("YWM_LIMIT", "100"))
    offset = int(os.getenv("YWM_OFFSET", "0"))
    out_dir = Path(os.getenv("YWM_OUTPUT_DIR", "out")).expanduser()

    if not token:
        raise SystemExit("Не задан YWM_OAUTH_TOKEN в .env")
    if not host_domain:
        raise SystemExit("Не задан YWM_HOST_DOMAIN в .env (домен без схемы)")

    headers = make_headers(token)
    user_id = get_user_id(headers)
    host_id = find_host_id(headers, user_id, host_domain)

    if not host_id:
        raise SystemExit(f"Хост с доменом '{host_domain}' не найден в вашем аккаунте Вебмастера.")

    print(f"Выгружаем внешние ссылки: user_id={user_id} host_id={host_id} limit={limit} offset={offset}")

    rows = iter_external_links(headers, user_id, host_id, limit=limit, offset=offset)
    write_xlsx(rows, out_dir, host_domain)


if __name__ == '__main__':
    main()