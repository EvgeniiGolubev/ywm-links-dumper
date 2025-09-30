from datetime import datetime
from pathlib import Path
from typing import Iterable, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from src.constants import XLSX_FIELDS, BAD_KEYWORDS

def write_xlsx(rows: Iterable[Dict], out_dir: str, host_domain: str) -> str:
    file_name = __make_file_name(host_domain)
    out_path = Path(out_dir).expanduser() / file_name
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "External Links"

    ws.append(XLSX_FIELDS)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for row in rows:
        values = [row.get(k, "") for k in XLSX_FIELDS]
        ws.append(values)

    # подсветка "плохих" ссылок
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        source_url = str(row[0].value).lower()
        if __is_bad_url(source_url):
            for cell in row:
                cell.fill = red_fill

    wb.save(out_path)

    return f"Файл сохранён: {out_path.resolve()}"

def __make_file_name(host_domain):
    today = datetime.today().strftime("%Y-%m-%d")
    safe_domain = host_domain.replace(":", "_")
    return f"{safe_domain}_links_{today}.xlsx"

def __is_bad_url(url: str) -> bool:
    s = (url or "").lower()
    return any(k in s for k in BAD_KEYWORDS)